import pandas as pd
from django.shortcuts import render
import os
from django.http import JsonResponse, HttpResponse
import openpyxl
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime
def submit_form(request):
    years = range(2000, datetime.now().year+1)
    if request.method == 'POST':
        movie_name = request.POST.get('movie_name')
        director_name = request.POST.get('director_name')
        year = request.POST.get('year')
        language = request.POST.get('language')
        rating = request.POST.get('rating')

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        DATA_DIR = os.path.join(BASE_DIR, 'static')

        file_path = os.path.join(DATA_DIR, 'movie_data.xlsx')

        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()

        worksheet = workbook.active

        next_row = worksheet.max_row + 1

        worksheet.cell(row=next_row, column=1, value=movie_name)
        worksheet.cell(row=next_row, column=2, value=director_name)
        worksheet.cell(row=next_row, column=3, value=year)
        worksheet.cell(row=next_row, column=4, value=language)
        worksheet.cell(row=next_row, column=5, value=rating)

        workbook.save(file_path)

        return render(request, 'add_movies.html',{'years': years})

    return render(request, 'add_movies.html',{'years': years})


def index(request):
    years = range(2000, datetime.now().year+1)
    return render(request, 'index.html', {'years': years})

def filter_movies(request):
    years = range(2000, datetime.now().year+1)
    return render(request, 'filter_movies.html')

def add_movies(request):
    years = range(2000, datetime.now().year+1)
    return render(request, 'add_movies.html', {'years': years})

def search_movie_by_language(request):
    years = range(2000, datetime.now().year+1)
    return render(request, 'search_movie_by_language.html')



def movie_list(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DATA_DIR = os.path.join(BASE_DIR, 'static')

    excel_file_path = os.path.join(DATA_DIR, 'movie_data.xlsx')

    df = pd.read_excel(excel_file_path)
    movies = df.to_dict('records')

    return JsonResponse(movies, safe=False)


@csrf_exempt
def delete_movie(request):
    if request.method == 'DELETE':
        try:
            movie_name = request.GET.get('movie_name')
            if movie_name:
                BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                DATA_DIR = os.path.join(BASE_DIR, 'static')

                excel_file_path = os.path.join(DATA_DIR, 'movie_data.xlsx')
                df = pd.read_excel(excel_file_path)

                df = df[df['Movie Name'] != movie_name]

                df.to_excel(excel_file_path, index=False)

                return JsonResponse({'message': 'Movie deleted successfully'}, status=200)
            else:
                return JsonResponse({'error': 'Movie name not provided'}, status=400)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

def get_movie_counts(request):
    if request.method == 'POST':
        selected_language = request.POST.get('language')
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        DATA_DIR = os.path.join(BASE_DIR, 'static')

        excel_file_path = os.path.join(DATA_DIR, 'movie_data.xlsx')
        df = pd.read_excel(excel_file_path)
        language_counts = df[df['Language'] == selected_language]['Movie Name'].count()

        return render(request, 'search_movie_by_language.html', {'language': selected_language, 'count': language_counts})
    return render(request, 'search_movie_by_language.html')



@csrf_exempt
def update_movie(request):
    if request.method == 'PUT':
        try:
            body = json.loads(request.body.decode('utf-8'))
            movie_name = body.get('movie_name')
            director_name = body.get('director_name')
            year = body.get('year')
            language = body.get('language')
            rating = body.get('rating')
            original_movie_name = body.get('original_movie_name')

            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            DATA_DIR = os.path.join(BASE_DIR, 'static')

            excel_file_path = os.path.join(DATA_DIR, 'movie_data.xlsx')
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value == original_movie_name:
                    row[0].value = movie_name
                    row[1].value = director_name
                    row[2].value = year
                    row[3].value = language
                    row[4].value = rating
                    break

            wb.save(excel_file_path)
            return HttpResponse(status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)
